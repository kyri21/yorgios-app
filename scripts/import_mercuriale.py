#!/usr/bin/env python3
"""
Import Mercuriale 949 YORGIOS → Firestore collection `mercuriale`
Projet : cuisine-yorgios / DB ID : test

Usage :
  python3 scripts/import_mercuriale.py --dry-run   # aperçu sans écriture
  python3 scripts/import_mercuriale.py             # import réel
"""
import sys
import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

DRY_RUN   = '--dry-run' in sys.argv
SA_KEY    = 'cuisine-yorgios-firebase-adminsdk-fbsvc-1c759ed390.json'
XLSX_PATH = 'reference/data/Mercuriale 949 YORGIOS.xlsx'

def norm_categorie(c: str) -> str:
    """Normalise les variantes (DESSERT → DESSERTS)."""
    c = c.strip().upper()
    if c == 'DESSERT':
        return 'DESSERTS'
    return c

def main():
    # ── Lire le fichier Excel ───────────────────────────────────────────────────
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['MERCURIALE 949']

    produits = []
    seen = set()

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        categorie = row[0]
        name      = row[1]
        unite     = row[2]
        prix      = row[3]

        if not all([categorie, name, unite, prix]):
            continue

        name_clean = str(name).strip().upper()
        unite_clean = str(unite).strip().upper()

        # Dédupliquer sur (nom, unité)
        key = (name_clean, unite_clean)
        if key in seen:
            print(f'  [DOUBLON ignoré] {name_clean}')
            continue
        seen.add(key)

        produits.append({
            'name':         name_clean,
            'categorie':    norm_categorie(str(categorie)),
            'unite':        unite_clean,          # 'KG' ou 'PIECE'
            'prixUnitaire': float(prix),
            'active':       True,
        })

    print(f'\n{len(produits)} produits à importer dans `mercuriale`\n')

    if DRY_RUN:
        for p in produits:
            print(f"  {p['categorie']:22} | {p['unite']:6} | {p['prixUnitaire']:7.3f} € | {p['name']}")
        print('\n[DRY-RUN] Aucune écriture effectuée.')
        return

    # ── Firebase ────────────────────────────────────────────────────────────────
    cred = credentials.Certificate(SA_KEY)
    firebase_admin.initialize_app(cred)
    db = firestore.client(database_id='test')

    # Supprimer l'ancienne collection
    existing = list(db.collection('mercuriale').stream())
    if existing:
        print(f'Suppression de {len(existing)} anciens docs...')
        for d in existing:
            d.reference.delete()

    # Importer
    col = db.collection('mercuriale')
    for p in produits:
        col.add(p)
        print(f"  ✓ {p['categorie']:22} | {p['unite']:6} | {p['prixUnitaire']:7.3f} € | {p['name']}")

    print(f'\n✅ {len(produits)} produits importés dans `mercuriale` (DB test)')

if __name__ == '__main__':
    main()
