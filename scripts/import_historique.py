"""
Import historique Excel → Firestore (DB: test, projet: cuisine-yorgios)

Usage:
    pip install firebase-admin openpyxl
    python scripts/import_historique.py [--dry-run]

Fichiers lus:
    reference/data/releve_temperature.xlsx  → temperatures
    reference/data/Hygiene.xlsx             → hygiene_corner
    reference/data/europoseidon_liaison.xlsx→ livraisons + objectifs_ca
"""

import sys
import datetime
import re
import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

DRY_RUN = '--dry-run' in sys.argv
SA_KEY   = 'cuisine-yorgios-firebase-adminsdk-fbsvc-1c759ed390.json'
DB_NAME  = 'test'

# ── Mapping frigos Excel → app ────────────────────────────────────────────────
# Frigo 1, 2, 3 = même frigo physique → FRIGO_3P
FRIGO_MAP = {
    'Frigo 1':   'FRIGO_3P',
    'Frigo 2':   'FRIGO_3P',
    'Frigo 3':   'FRIGO_3P',
    'Vitrine 1': 'VITRINE_1',
    'Vitrine 2': 'VITRINE_2',
    'Vitrine 3': 'VITRINE_3',
    'Grand frigo': 'GRAND_FRIGO',
    'Grand Frigo': 'GRAND_FRIGO',
}
FRIGO_NAMES = {
    'FRIGO_3P':    'Frigo 3 portes',
    'VITRINE_1':   'Vitrine 1',
    'VITRINE_2':   'Vitrine 2',
    'VITRINE_3':   'Vitrine 3',
    'GRAND_FRIGO': 'Grand frigo',
}

# ── Mapping hygiène Excel → app item IDs ─────────────────────────────────────
HYGIENE_QUOTIDIEN_MAP = {
    'Plats de service':                 'plats_service',
    'Intérieur vitrines libre service': 'int_vitrines',
    'Ustensiles':                       'ustensiles',
    'Meuble de vente':                  'meuble_vente',
    'Comptoir / balance':               'comptoir_balance',
    'Micro-ondes':                      'micro_ondes',
    'Évier / Distributeur papier':      'evier_papier',
    'Étiquettes':                       'etiquettes',
    'Plan de travail':                  'plan_travail',
    'Extérieur placards rangement':     'ext_placards',
    'Extérieur frigo':                  'ext_frigo',
    'Poubelle':                         'poubelle',
    'Vitres':                           'vitres',
}
HYGIENE_HEBDO_MAP = {
    'Intérieur frigos':            'int_frigos',
    'Étagères porte matériels':    'etageres_materiels',
    'Support rouleau papier':      'support_papier',
    'Placard hygiène':             'placard_hygiene',
    'Machine à Glaçons':           'machine_glacon',
}
HYGIENE_MENSUEL_MAP = {
    'Placard rangement': 'placard_rangement',
}

ALERT_MIN = -2
ALERT_MAX = 4

def init_firebase():
    cred = credentials.Certificate(SA_KEY)
    firebase_admin.initialize_app(cred, {'projectId': 'cuisine-yorgios'})
    return firestore.client(database_id=DB_NAME)

def parse_temp(val) -> float | None:
    """Parse une valeur de cellule Excel en float, retourne None si invalide."""
    if val is None or val == '':
        return None
    s = str(val).strip().replace(',', '.').replace('+', '').replace(' ', '')
    # Cas "Df", "df" = dégivrage → on ignore
    if re.fullmatch(r'[Dd][Ff].*', s):
        return None
    # Cas "1..6" → "1.6"
    s = re.sub(r'\.{2,}', '.', s)
    try:
        return float(s)
    except ValueError:
        return None

def compute_status(t: float) -> str:
    if t < ALERT_MIN or t > ALERT_MAX:
        return 'ALERTE'
    return 'OK'

def iso_week_to_monday(year: int, week: int) -> datetime.date:
    """Retourne le lundi de la semaine ISO donnée."""
    jan4 = datetime.date(year, 1, 4)
    start_w1 = jan4 - datetime.timedelta(days=jan4.weekday())
    return start_w1 + datetime.timedelta(weeks=week - 1)

def is_checked(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    return s in ('✅', '✓', '1', 'True', 'true', 'oui', 'Oui', 'x', 'X', 'OK')

def batch_set(db, collection: str, doc_id: str, data: dict, written: list):
    if DRY_RUN:
        print(f'  [DRY] {collection}/{doc_id}')
        return
    ref = db.collection(collection).document(doc_id)
    # Ne pas écraser un doc existant
    existing = ref.get()
    if existing.exists:
        print(f'  [SKIP] {collection}/{doc_id} existe déjà')
        return
    ref.set(data)
    written.append(f'{collection}/{doc_id}')

# ─────────────────────────────────────────────────────────────────────────────
# 1. TEMPÉRATURES
# ─────────────────────────────────────────────────────────────────────────────

def import_temperatures(db):
    print('\n=== Températures ===')
    wb = openpyxl.load_workbook('reference/data/releve_temperature.xlsx', read_only=True, data_only=True)
    written = []
    skipped_frigos = set()

    for sheet_name in wb.sheetnames:
        # "Semaine 16 2025" → week=16, year=2025
        m = re.match(r'Semaine\s+(\d+)\s+(\d{4})', sheet_name, re.IGNORECASE)
        if not m:
            print(f'  [SKIP] onglet ignoré: {sheet_name}')
            continue
        week, year = int(m.group(1)), int(m.group(2))
        monday = iso_week_to_monday(year, week)

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Ligne 0 = en-têtes: Frigo/Vitrine | Lundi Matin | Lundi Soir | ...
        headers = rows[0]
        # Construire l'index des colonnes : (jour_index 0-6, session 'matin'|'soir')
        col_info = {}  # col_index → (day_offset, session)
        for ci, h in enumerate(headers):
            if h is None or ci == 0:
                continue
            hs = str(h).strip().lower()
            for day_offset, day_name in enumerate(['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi', 'dimanche']):
                # handle typo "mecredi"
                if day_name.startswith(hs[:4]) or hs.startswith(day_name[:4]):
                    if 'matin' in hs:
                        col_info[ci] = (day_offset, 'matin')
                    elif 'soir' in hs:
                        col_info[ci] = (day_offset, 'soir')
                    break

        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            frigo_label = str(row[0]).strip()
            # Deduplicate: Frigo 1/2/3 → FRIGO_3P, skip 2 et 3
            if frigo_label in ('Frigo 2', 'Frigo 3'):
                skipped_frigos.add(frigo_label)
                continue
            frigo_id = FRIGO_MAP.get(frigo_label)
            if not frigo_id:
                print(f'  [WARN] frigo inconnu: {frigo_label}')
                continue

            for ci, (day_offset, session) in col_info.items():
                if ci >= len(row):
                    continue
                t = parse_temp(row[ci])
                if t is None:
                    continue
                date = monday + datetime.timedelta(days=day_offset)
                date_str = date.isoformat()
                doc_id = f'{date_str}_{frigo_id}_{session}'
                data = {
                    'date': date_str,
                    'fridgeId': frigo_id,
                    'fridgeName': FRIGO_NAMES[frigo_id],
                    'session': session,
                    'tempC': t,
                    'status': compute_status(t),
                    'alertMin': ALERT_MIN,
                    'alertMax': ALERT_MAX,
                    'source': 'import_excel',
                    'createdBy': 'import',
                    'createdAt': firestore.SERVER_TIMESTAMP,
                }
                batch_set(db, 'temperatures', doc_id, data, written)

    wb.close()
    if skipped_frigos:
        print(f'  Frigos dupliqués ignorés (fusionnés dans FRIGO_3P): {skipped_frigos}')
    print(f'  → {len(written)} docs écrits')

# ─────────────────────────────────────────────────────────────────────────────
# 2. HYGIÈNE
# ─────────────────────────────────────────────────────────────────────────────

def iso_week_str(date_obj: datetime.date) -> str:
    iso = date_obj.isocalendar()
    return f'{iso[0]}-W{str(iso[1]).zfill(2)}'

def parse_date_cell(val) -> datetime.date | None:
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    s = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def import_hygiene(db):
    print('\n=== Hygiène ===')
    wb = openpyxl.load_workbook('reference/data/Hygiene.xlsx', read_only=True, data_only=True)
    written = []

    # ── Quotidien ──
    ws = wb['Quotidien']
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        date = parse_date_cell(row[0])
        if date is None:
            continue
        date_str = date.isoformat()
        items: dict[str, bool] = {}
        for ci, h in enumerate(headers[1:], start=1):
            if h is None or ci >= len(row):
                continue
            item_id = HYGIENE_QUOTIDIEN_MAP.get(str(h).strip())
            if item_id:
                items[item_id] = is_checked(row[ci])
        doc_id = f'{date_str}_quotidien'
        batch_set(db, 'hygiene_corner', doc_id, {
            'items': items, 'source': 'import_excel',
            'createdBy': 'import', 'createdAt': firestore.SERVER_TIMESTAMP,
        }, written)

    # ── Hebdomadaire ──
    ws = wb['Hebdomadaire']
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        date = parse_date_cell(row[0])
        if date is None:
            continue
        week_str = iso_week_str(date)
        items: dict[str, bool] = {}
        for ci, h in enumerate(headers[1:], start=1):
            if h is None or ci >= len(row):
                continue
            item_id = HYGIENE_HEBDO_MAP.get(str(h).strip())
            if item_id:
                items[item_id] = is_checked(row[ci])
        doc_id = f'{week_str}_hebdo'
        batch_set(db, 'hygiene_corner', doc_id, {
            'items': items, 'source': 'import_excel',
            'createdBy': 'import', 'createdAt': firestore.SERVER_TIMESTAMP,
        }, written)

    # ── Mensuel ──
    ws = wb['Mensuel']
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        date = parse_date_cell(row[0])
        if date is None:
            continue
        month_str = f'{date.year}-{str(date.month).zfill(2)}'
        items: dict[str, bool] = {}
        for ci, h in enumerate(headers[1:], start=1):
            if h is None or ci >= len(row):
                continue
            item_id = HYGIENE_MENSUEL_MAP.get(str(h).strip())
            if item_id:
                items[item_id] = is_checked(row[ci])
        doc_id = f'{month_str}_mensuel'
        batch_set(db, 'hygiene_corner', doc_id, {
            'items': items, 'source': 'import_excel',
            'createdBy': 'import', 'createdAt': firestore.SERVER_TIMESTAMP,
        }, written)

    wb.close()
    print(f'  → {len(written)} docs écrits')

# ─────────────────────────────────────────────────────────────────────────────
# 3. LIVRAISONS (températures de livraison)
# ─────────────────────────────────────────────────────────────────────────────

CAT_MAP = {
    'viande hachée':     'VIANDE_HACHEE',
    'viande':            'VIANDE',
    'plat cuisiné':      'PLAT_CUISINE',
    'plats cuisinés':    'PLAT_CUISINE',
    'lait':              'LAIT',
    'pâtisserie':        'PATISSERIE',
    'légume':            'LEGUME',
    'légumes':           'LEGUME',
}

def import_livraisons(db):
    print('\n=== Livraisons température ===')
    wb = openpyxl.load_workbook('reference/data/europoseidon_liaison.xlsx', read_only=True, data_only=True)
    written = []

    ws = wb['Livraison Température']
    rows = list(ws.iter_rows(values_only=True))
    # Headers: Produit | Temp départ | Horodatage départ | Temp réception | Dénomination GEP | Résultat | Lien photo
    for i, row in enumerate(rows[1:], start=2):
        if not row or row[0] is None:
            continue
        produit = str(row[0]).strip() if row[0] else ''
        depart_temp = parse_temp(row[1])
        horodatage = row[2]
        reception_temp = parse_temp(row[3])
        cat_raw = str(row[4]).strip().lower() if row[4] else ''
        resultat_raw = str(row[5]).strip() if row[5] else ''

        category = CAT_MAP.get(cat_raw, 'PLAT_CUISINE')
        result = 'ACCEPTE' if '✅' in resultat_raw or 'Accepté' in resultat_raw else 'REFUSE'

        if isinstance(horodatage, datetime.datetime):
            depart_at = horodatage
        else:
            depart_at = datetime.datetime.now()

        doc_id = f'import_{depart_at.strftime("%Y%m%d_%H%M%S")}_{i}'
        lot_code = f'{depart_at.strftime("%d%m%Y")}-IMP-{produit[:6].upper().replace(" ", "")}'

        data = {
            'lotId': None,
            'lotCode': lot_code,
            'productId': None,
            'productName': produit,
            'category': category,
            'departTempC': depart_temp,
            'departAt': depart_at,
            'departBy': 'import',
            'receptionTempC': reception_temp,
            'receptionAt': depart_at,
            'receptionBy': 'import',
            'result': result,
            'ruleMaxTol': None,
            'isManual': True,
            'source': 'import_excel',
            'createdAt': depart_at,
        }
        batch_set(db, 'livraisons', doc_id, data, written)

    wb.close()
    print(f'  → {len(written)} docs écrits')

# ─────────────────────────────────────────────────────────────────────────────
# 4. OBJECTIFS CA
# ─────────────────────────────────────────────────────────────────────────────

MONTH_FR = {
    'janvier': 1, 'février': 2, 'fevrier': 2, 'mars': 3, 'avril': 4,
    'mai': 5, 'juin': 6, 'juillet': 7, 'août': 8, 'aout': 8,
    'septembre': 9, 'octobre': 10, 'novembre': 11, 'décembre': 12, 'decembre': 12,
}

def import_objectifs_ca(db):
    print('\n=== Objectifs CA ===')
    wb = openpyxl.load_workbook('reference/data/europoseidon_liaison.xlsx', read_only=True, data_only=True)
    written = []

    ws = wb['Objectifs']
    rows = list(ws.iter_rows(values_only=True))
    year = datetime.date.today().year  # les objectifs sont pour l'année en cours

    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        month_name = str(row[0]).strip().lower()
        month_num = MONTH_FR.get(month_name)
        if month_num is None:
            continue
        objectif = row[1]
        resultat_raw = row[2]
        if objectif is None:
            continue
        try:
            objectif = float(str(objectif).replace(',', '.').replace(' ', ''))
        except ValueError:
            continue
        resultat = None
        if resultat_raw is not None:
            try:
                resultat = float(str(resultat_raw).replace(',', '.').replace(' ', ''))
            except ValueError:
                pass

        doc_id = f'{year}-{str(month_num).zfill(2)}'
        data = {'objectif': objectif, 'source': 'import_excel'}
        if resultat is not None:
            data['resultat'] = resultat

        batch_set(db, 'objectifs_ca', doc_id, data, written)

    wb.close()
    print(f'  → {len(written)} docs écrits')

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import os
    os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    if DRY_RUN:
        print('=== MODE DRY RUN — aucune écriture Firestore ===')

    print(f'Connexion Firestore ({DB_NAME})…')
    db = init_firebase()
    print('Connecté.')

    import_temperatures(db)
    import_hygiene(db)
    import_livraisons(db)
    import_objectifs_ca(db)

    print('\n=== Import terminé ===')
