"""
Import Planning Excel → Firestore (DB: test, projet: cuisine-yorgios)

Usage:
    pip install firebase-admin openpyxl
    python scripts/import_planning.py [--dry-run]
    python scripts/import_planning.py --file /chemin/vers/planning.xlsx

Structure Firestore générée :
    planningWeeks/{YYYY-MM-DD}                      ← doc semaine (lundi)
    planningWeeks/{YYYY-MM-DD}/days/{0..6}          ← un doc par jour
        .hours = {"8": [empId, ...], "9": [...], ...}

Mapping initiales Excel → noms (à adapter si besoin) :
    D → Arthur        S → Sébastien    A → Alexandre
    E → Elena         K → Markella     Y → Layal
    M → Mellina       W → Wahib
    N, X → cherchés automatiquement dans Firestore par le champ 'initials'
"""

import sys
import datetime
import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

DRY_RUN   = '--dry-run' in sys.argv
SA_KEY    = 'cuisine-yorgios-firebase-adminsdk-fbsvc-1c759ed390.json'
DB_NAME   = 'test'

# Fichier par défaut — override avec --file
EXCEL_FILE = '/home/demis/Téléchargements/Planning temporaire du 8mars.xlsx'
for i, arg in enumerate(sys.argv):
    if arg == '--file' and i + 1 < len(sys.argv):
        EXCEL_FILE = sys.argv[i + 1]

# Mapping initiales Excel → prénom employé (recherche ensuite dans Firestore)
INITIAL_TO_NAME = {
    'D': 'Arthur',
    'S': 'Sébastien',
    'A': 'Alexandre',
    'E': 'Elena',
    'K': 'Markella',
    'Y': 'Layal',
    'M': 'Mellina',
    'N': 'Mellina',
    'W': 'Wahib',
    'X': 'Wahib',
}


def parse_initials(cell_value):
    """'D+A+S' → ['D', 'A', 'S']"""
    if not cell_value:
        return []
    return [i.strip().upper() for i in str(cell_value).split('+') if i.strip()]


def main():
    # ── Init Firebase ────────────────────────────────────────────────────────
    cred = credentials.Certificate(SA_KEY)
    firebase_admin.initialize_app(cred)
    db_client = firestore.client(database_id=DB_NAME)

    # ── Charger les employés depuis Firestore ────────────────────────────────
    print('Chargement des employés depuis Firestore...')
    emp_docs = db_client.collection('employees').stream()

    emp_by_name_lower = {}   # "arthur"     → emp_id
    emp_by_initial    = {}   # "D" (champ)  → emp_id

    for doc in emp_docs:
        data = doc.to_dict()
        name = data.get('name', '')
        initials_field = (data.get('initials') or '').strip().upper()
        emp_id = doc.id

        emp_by_name_lower[name.lower()] = emp_id
        first = name.split()[0].lower()
        emp_by_name_lower[first] = emp_id

        if initials_field:
            emp_by_initial[initials_field] = emp_id

        active = data.get('active', True)
        print(f"  {'✓' if active else '✗'} {name}  (id={emp_id}, initials='{initials_field}')")

    def resolve_initial(initial):
        """Résout une initiale Excel en emp_id Firestore. Retourne None si non trouvé."""
        initial = initial.strip().upper()
        # 1) Par le mapping nom
        name = INITIAL_TO_NAME.get(initial)
        if name:
            emp_id = emp_by_name_lower.get(name.lower())
            if emp_id:
                return emp_id
            first = name.split()[0].lower()
            emp_id = emp_by_name_lower.get(first)
            if emp_id:
                return emp_id
        # 2) Par le champ 'initials' dans Firestore (pour N, X, etc.)
        if initial in emp_by_initial:
            return emp_by_initial[initial]
        return None

    # ── Lire l'Excel ─────────────────────────────────────────────────────────
    print(f'\nLecture : {EXCEL_FILE}')
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    print(f'{len(wb.sheetnames)} feuilles trouvées.')

    weeks_ok  = 0
    weeks_skip = 0
    unknown_initials = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Ligne 2, colonne B = date du lundi
        date_cell = ws.cell(row=2, column=2).value
        if not isinstance(date_cell, datetime.datetime):
            print(f'  [SKIP] {sheet_name} : date invalide ({date_cell!r})')
            weeks_skip += 1
            continue

        monday   = date_cell.date()
        week_id  = monday.strftime('%Y-%m-%d')

        # Construire la map heures par jour
        # Colonnes : B=2 (Lun/jour 0) … H=8 (Dim/jour 6)
        # Lignes   : 5-17 = heures 8-20
        days_hours = {}
        for day_idx in range(7):
            col = day_idx + 2   # B=2 … H=8
            hours_map = {}
            for row in range(5, 18):   # rows 5-17
                hour = row - 5 + 8     # 8 .. 20
                cell_val = ws.cell(row=row, column=col).value
                initials = parse_initials(cell_val)
                emp_ids = []
                for ini in initials:
                    eid = resolve_initial(ini)
                    if eid:
                        emp_ids.append(eid)
                    elif ini:
                        unknown_initials.add(ini)
                hours_map[str(hour)] = emp_ids
            days_hours[day_idx] = hours_map

        # Aperçu
        filled_days = sum(
            1 for hm in days_hours.values()
            if any(ids for ids in hm.values())
        )
        print(f'  {"[DRY]" if DRY_RUN else "→"} {week_id}  ({sheet_name})  — {filled_days}/7 jours remplis')

        if DRY_RUN:
            weeks_ok += 1
            continue

        # ── Écriture Firestore ──────────────────────────────────────────────
        week_ref = db_client.collection('planningWeeks').document(week_id)
        week_ref.set({
            'weekId':      week_id,
            'mondayDate':  week_id,
            'updatedAt':   firestore.SERVER_TIMESTAMP,
            'updatedBy':   'import_planning',
            'locked':      False,
        }, merge=True)

        for day_idx, hours_map in days_hours.items():
            day_ref = week_ref.collection('days').document(str(day_idx))
            day_ref.set({
                'dayIndex':  day_idx,
                'hours':     hours_map,
                'updatedAt': firestore.SERVER_TIMESTAMP,
                'updatedBy': 'import_planning',
            })

        weeks_ok += 1

    # ── Résumé ───────────────────────────────────────────────────────────────
    print(f'\n{"=" * 40}')
    print(f'Semaines importées : {weeks_ok}')
    print(f'Semaines ignorées  : {weeks_skip}')
    if unknown_initials:
        print(f'\nInitiales non résolues (ignorées) : {sorted(unknown_initials)}')
        print('  → Vérifie que ces employés existent dans Firestore (collection "employees")')
        print('    avec le champ "initials" correspondant, ou ajoute-les dans INITIAL_TO_NAME.')
    print('Done ✓')


if __name__ == '__main__':
    main()
